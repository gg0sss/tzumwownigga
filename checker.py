import requests
import json
import os
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook

BOT_TOKEN = os.environ["BOT_TOKEN"]
CHAT_ID = os.environ["CHAT_ID"]
TG_API = f"https://api.telegram.org/bot{BOT_TOKEN}"
DB_FILE = "products.json"
EXCEL_FILE = "sales_history.xlsx"

CATEGORIES = [
    "https://collect.tsum.ru/women/catalog/povsednevnye-sumki-82",
    "https://collect.tsum.ru/women/catalog/riukzaki-i-poiasnye-sumki-87",
    "https://collect.tsum.ru/women/catalog/dorozhnye-i-sportivnye-sumki-93",
    "https://collect.tsum.ru/women/catalog/klatchi-i-vechernie-sumki-90",
    "https://collect.tsum.ru/men/catalog/riukzaki-i-poiasnye-sumki-246",
    "https://collect.tsum.ru/men/catalog/povsednevnye-sumki-238",
    "https://collect.tsum.ru/men/catalog/dorozhnye-i-sportivnye-sumki-249"
]

def send(msg):
    try:
        requests.post(f"{TG_API}/sendMessage", json={"chat_id": CHAT_ID, "text": msg})
        
        chat_id_2 = os.environ.get("CHAT_ID_2")
        if chat_id_2:
            requests.post(f"{TG_API}/sendMessage", json={"chat_id": chat_id_2, "text": msg})
    except Exception as e:
        print(f"Ошибка отправки: {e}")

def init_excel():
    """Создаёт Excel файл если его нет"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "История продаж"
        ws.append(["Дата продажи", "Бренд", "Цена", "Месяц выставления", "Ссылка"])
        wb.save(EXCEL_FILE)
        print("✅ Создан файл sales_history.xlsx")

def add_to_excel(brand, price, listing_date, url):
    """Добавляет проданный товар в Excel"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        sale_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([sale_date, brand, price, listing_date, url])
        
        wb.save(EXCEL_FILE)
        print(f"  📊 Добавлено в Excel")
    except Exception as e:
        print(f"Ошибка записи в Excel: {e}")

def estimate_listing_date(item_url):
    """Определяет примерный месяц размещения по номеру ITEM"""
    try:
        item_id = item_url.split("/item/")[1].split("/")[0]
        num = int(item_id.replace("ITEM", ""))
        
        if num >= 378324: return "декабрь 2025"
        elif num >= 375363: return "ноябрь 2025"
        elif num >= 374536: return "октябрь 2025"
        elif num >= 366646: return "август 2025"
        elif num >= 362999: return "июнь 2025"
        elif num >= 350905: return "май 2025"
        elif num >= 332922: return "начало 2025"
        elif num >= 305982: return "конец 2024"
        elif num >= 221563: return "2023-2024"
        else: return "очень давно"
    except:
        return "неизвестно"

def check_product_page(driver, url):
    try:
        driver.get(url)
        time.sleep(3)
        
        try:
            driver.find_element(By.CSS_SELECTOR, "p[class*='noExists']")
            return "sold"
        except:
            pass
        
        try:
            driver.find_element(By.CSS_SELECTOR, "p[class*='reserved']")
            return "reserved"
        except:
            pass
        
        return "available"
    except Exception as e:
        print(f"Ошибка проверки {url}: {e}")
        return "unknown"

# Инициализируем Excel
init_excel()

if os.path.exists(DB_FILE):
    with open(DB_FILE, "r", encoding="utf-8") as f:
        old_products = json.load(f)
else:
    old_products = {}

new_products = {}

chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--window-size=1920,1080")
chrome_options.add_argument("user-agent=Mozilla/5.0")

try:
    send("Ищу дорогую ненужную хуйню🥶")
    driver = webdriver.Chrome(options=chrome_options)
    
    for category_url in CATEGORIES:
        print(f"\nПарсинг: {category_url}")
        driver.get(category_url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='/item/ITEM']")))
        
        attempts = 0
        while attempts < 200:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            try:
                button = driver.find_element(By.XPATH, "//p[contains(text(), 'Показать больше товаров')]")
                driver.execute_script("arguments[0].click();", button)
                time.sleep(3)
            except:
                break
            attempts += 1
        
        cards = driver.find_elements(By.CSS_SELECTOR, "a[href*='/item/ITEM']")
        print(f"  Найдено: {len(cards)} товаров")
        
        for card in cards:
            try:
                url = card.get_attribute("href")
                if url in new_products:
                    continue
                
                try:
                    brand_img = card.find_element(By.CSS_SELECTOR, "img[data-brandlogo='true']")
                    brand_name = brand_img.get_attribute("alt") or "Товар"
                except:
                    brand_name = "Товар"
                
                try:
                    price_elem = card.find_element(By.CSS_SELECTOR, "span[class*='price']")
                    price_text = price_elem.text.strip()
                except:
                    price_text = "Цена неизвестна"
                
                new_products[url] = {
                    "title": brand_name,
                    "price": price_text,
                    "in_stock": True
                }
            except Exception as e:
                continue
    
    print(f"\n✅ Всего товаров: {len(new_products)}")
    
    sold_count = 0
    for old_url, old_data in old_products.items():
        if old_data["in_stock"] and old_url not in new_products:
            print(f"Проверяю: {old_url}")
            status = check_product_page(driver, old_url)
            
            if status == "sold":
                price_info = old_data.get('price', 'Цена неизвестна')
                listing_date = estimate_listing_date(old_url)
                
                # Добавляем в Excel
                add_to_excel(old_data['title'], price_info, listing_date, old_url)
                
                send(f"❌ ПРОДАНО\n\n{old_data['title']}\nЦена: {price_info}\nВыставлено: {listing_date}\n\n{old_url}")
                sold_count += 1
                print(f"  ✅ ПРОДАНО: {old_data['title']} за {price_info} (выставлено: {listing_date})")
            elif status == "reserved":
                print(f"  В резерве - игнорируем")
    
    driver.quit()
    
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(new_products, f, ensure_ascii=False, indent=2)
    
    send(f"✅ Нашел дэм🫨\nТоваров: {len(new_products)}\nПродано: {sold_count}")

except Exception as e:
    send(f"⚠️ Ошибка: {str(e)}")
    print(f"ERROR: {e}")
    try:
        driver.quit()
    except:
        pass
